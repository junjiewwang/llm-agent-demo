"""文件读取与搜索工具。

提供 5 种操作，通过 action 参数区分：
- list_allowed_dirs: 列出所有可访问的目录及权限
- list_dir: 列出目录内容（含文件大小、修改时间）
- find_files: 按 glob 模式查找文件名
- search_content: 在文件内容中搜索文本（类似 grep）
- read_file: 读取文件内容（支持分段读取）

所有操作都受 Sandbox 沙箱约束，确保路径安全。
支持多根目录白名单，路径可以是相对路径、绝对路径或 ~ 路径。
"""

import os
import re
import time
from pathlib import Path
from typing import Any, Dict, List

from src.tools.base_tool import BaseTool
from src.tools.filesystem.sandbox import Sandbox
from src.utils.logger import logger

# 已知的二进制/富格式文件扩展名，无法以文本方式有效读取
_BINARY_EXTENSIONS: set[str] = {
    # 压缩/归档
    ".zip", ".tar", ".gz", ".bz2", ".7z", ".rar", ".xz",
    # 图片
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".ico", ".webp", ".svg", ".tiff",
    # 音视频
    ".mp3", ".mp4", ".avi", ".mov", ".mkv", ".wav", ".flac", ".ogg",
    # 可执行/字节码
    ".exe", ".dll", ".so", ".dylib", ".pyc", ".pyo", ".class", ".o",
    # 数据库/二进制数据
    ".sqlite", ".db", ".bin", ".dat", ".pkl", ".pickle",
    # PDF
    ".pdf",
    # 字体
    ".ttf", ".otf", ".woff", ".woff2",
    # Office 文档（不含 Excel，Excel 已支持解析）
    ".docx", ".doc", ".pptx", ".ppt",
}

# 可通过专用解析器读取的富格式文件扩展名
_EXCEL_EXTENSIONS: set[str] = {".xlsx", ".xls"}


def _is_binary_file(file_path: Path) -> bool:
    """判断文件是否为已知的二进制/富格式文件。"""
    return file_path.suffix.lower() in _BINARY_EXTENSIONS


def _binary_file_hint(file_path: Path) -> str:
    """为二进制文件生成友好的提示信息。"""
    suffix = file_path.suffix.lower()
    hints = {
        ".docx": "Word 文档",
        ".doc": "Word 文档（旧版）",
        ".pptx": "PowerPoint 演示文稿",
        ".ppt": "PowerPoint 演示文稿（旧版）",
        ".pdf": "PDF 文档",
        ".zip": "ZIP 压缩包",
        ".tar": "TAR 归档",
        ".gz": "GZip 压缩文件",
    }
    file_type = hints.get(suffix, f"{suffix} 二进制文件")
    return (
        f"⚠️ 无法以文本方式读取: {file_path.name}\n"
        f"文件类型: {file_type}\n"
        f"建议: 如果是 PDF，请通过知识库上传功能导入。"
    )


def _format_size(size: int) -> str:
    """将字节数格式化为人类可读的大小。"""
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024:
            return f"{size:.0f}{unit}" if unit == "B" else f"{size:.1f}{unit}"
        size /= 1024
    return f"{size:.1f}TB"


def _format_time(ts: float) -> str:
    """将时间戳格式化为可读字符串。"""
    return time.strftime("%Y-%m-%d %H:%M", time.localtime(ts))


class FileReaderTool(BaseTool):
    """文件读取与搜索工具。

    通过 Sandbox 限制所有操作在安全目录内，
    支持目录浏览、文件查找、内容搜索、文件读取。
    """

    def __init__(self, sandbox: Sandbox):
        self._sandbox = sandbox

    @property
    def name(self) -> str:
        return "file_reader"

    @property
    def description(self) -> str:
        return (
            "读取和搜索本地文件系统。支持 5 种操作（通过 action 参数指定）：\n"
            "1. list_allowed_dirs: 列出所有可访问的目录及其权限（不需要 path 参数）\n"
            "2. list_dir: 列出目录内容（文件名、大小、修改时间）\n"
            "3. find_files: 按文件名模式查找（如 *.py、test_*）\n"
            "4. search_content: 在文件内容中搜索文本（类似 grep）\n"
            "5. read_file: 读取文件内容（支持指定起始行和行数，支持 .xlsx/.xls Excel 文件自动解析为表格）\n"
            "适用场景：需要浏览目录结构、查找特定文件、搜索代码中的关键词、阅读文件内容时使用。\n"
            "不适用：修改文件请使用 file_writer 工具。\n"
            "提示：可以先用 list_allowed_dirs 查看可访问哪些目录。路径支持相对路径、绝对路径和 ~ 路径。\n"
            f"限制：单文件最大 {_format_size(self._sandbox.max_file_size)}。"
        )

    @property
    def parameters(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["list_allowed_dirs", "list_dir", "find_files", "search_content", "read_file"],
                    "description": "操作类型",
                },
                "path": {
                    "type": "string",
                    "description": (
                        "目标路径。支持相对路径（基于默认工作目录）、绝对路径、~ 路径。"
                        "list_allowed_dirs: 不需要此参数；"
                        "list_dir/find_files/search_content: 目标目录；read_file: 文件路径"
                    ),
                },
                "pattern": {
                    "type": "string",
                    "description": (
                        "find_files: glob 模式（如 *.py、**/test_*.py）；"
                        "search_content: 搜索文本或正则表达式"
                    ),
                },
                "max_depth": {
                    "type": "integer",
                    "description": "搜索最大深度，默认 3",
                },
                "offset": {
                    "type": "integer",
                    "description": "read_file: 起始行号（从 1 开始），默认从头读取",
                },
                "limit": {
                    "type": "integer",
                    "description": "read_file: 读取行数，默认读取全部",
                },
            },
            "required": ["action"],
        }

    def execute(self, action: str, path: str = ".", **kwargs) -> str:
        """根据 action 分发到具体操作。"""
        dispatch = {
            "list_allowed_dirs": self._list_allowed_dirs,
            "list_dir": self._list_dir,
            "find_files": self._find_files,
            "search_content": self._search_content,
            "read_file": self._read_file,
        }

        handler = dispatch.get(action)
        if not handler:
            return f"未知操作: {action}。支持的操作: {list(dispatch.keys())}"

        try:
            return handler(path, **kwargs)
        except (PermissionError, FileNotFoundError, ValueError) as e:
            return f"操作失败: {e}"
        except Exception as e:
            logger.error("file_reader.{} 异常: {}", action, e)
            return f"操作异常: {e}"

    # ── 具体操作 ──

    def _list_allowed_dirs(self, path: str = ".", **kwargs) -> str:
        """列出所有可访问的目录及权限。"""
        return self._sandbox.list_allowed_dirs()

    def _list_dir(self, path: str, max_depth: int = 3, **kwargs) -> str:
        """列出目录内容，含文件大小和修改时间。"""
        dir_path = self._sandbox.validate_dir(path)
        max_depth = min(max_depth, self._sandbox.max_depth)

        lines: List[str] = [f"📁 {self._sandbox.relative_to_root(dir_path)}/"]
        self._walk_dir(dir_path, lines, prefix="", depth=0, max_depth=max_depth)

        if len(lines) == 1:
            lines.append("  (空目录)")

        return "\n".join(lines)

    def _walk_dir(
        self, dir_path: Path, lines: List[str],
        prefix: str, depth: int, max_depth: int,
    ) -> None:
        """递归遍历目录，构建树形输出。"""
        if depth >= max_depth:
            return

        try:
            entries = sorted(dir_path.iterdir(), key=lambda e: (not e.is_dir(), e.name.lower()))
        except PermissionError:
            lines.append(f"{prefix}  ⚠️ 权限不足")
            return

        # 限制结果数量
        count = 0
        for entry in entries:
            if self._sandbox.is_excluded(entry):
                continue

            count += 1
            if count > self._sandbox.max_results:
                lines.append(f"{prefix}  ... 还有更多项（已达上限 {self._sandbox.max_results}）")
                break

            rel = self._sandbox.relative_to_root(entry)
            if entry.is_dir():
                lines.append(f"{prefix}  📁 {entry.name}/")
                self._walk_dir(entry, lines, prefix=prefix + "  ", depth=depth + 1, max_depth=max_depth)
            else:
                stat = entry.stat()
                size = _format_size(stat.st_size)
                mtime = _format_time(stat.st_mtime)
                lines.append(f"{prefix}  📄 {entry.name}  ({size}, {mtime})")

    def _find_files(self, path: str, pattern: str = "*", max_depth: int = 3, **kwargs) -> str:
        """按 glob 模式查找文件。"""
        dir_path = self._sandbox.validate_dir(path)
        max_depth = min(max_depth, self._sandbox.max_depth)

        results: List[str] = []
        self._glob_search(dir_path, pattern, results, depth=0, max_depth=max_depth)

        if not results:
            return f"在 {self._sandbox.relative_to_root(dir_path)}/ 下未找到匹配 '{pattern}' 的文件"

        header = f"找到 {len(results)} 个匹配 '{pattern}' 的文件：\n"
        return header + "\n".join(results)

    def _glob_search(
        self, dir_path: Path, pattern: str,
        results: List[str], depth: int, max_depth: int,
    ) -> None:
        """递归 glob 搜索。"""
        if depth >= max_depth or len(results) >= self._sandbox.max_results:
            return

        try:
            entries = sorted(dir_path.iterdir(), key=lambda e: e.name.lower())
        except PermissionError:
            return

        for entry in entries:
            if self._sandbox.is_excluded(entry):
                continue

            if len(results) >= self._sandbox.max_results:
                results.append(f"... 结果已达上限 ({self._sandbox.max_results})")
                return

            if entry.is_file() and entry.match(pattern):
                rel = self._sandbox.relative_to_root(entry)
                stat = entry.stat()
                results.append(f"  {rel}  ({_format_size(stat.st_size)})")
            elif entry.is_dir():
                self._glob_search(entry, pattern, results, depth + 1, max_depth)

    def _search_content(self, path: str, pattern: str = "", max_depth: int = 3, **kwargs) -> str:
        """在文件内容中搜索文本（类似 grep）。"""
        if not pattern:
            return "search_content 需要提供 pattern 参数（搜索文本或正则表达式）"

        dir_path = self._sandbox.validate_dir(path)
        max_depth = min(max_depth, self._sandbox.max_depth)

        # 编译正则（如果是普通文本，re.escape 处理）
        try:
            regex = re.compile(pattern, re.IGNORECASE)
        except re.error:
            # 用户输入的不是合法正则，作为普通文本搜索
            regex = re.compile(re.escape(pattern), re.IGNORECASE)

        matches: List[str] = []
        self._grep_search(dir_path, regex, matches, depth=0, max_depth=max_depth)

        if not matches:
            return f"在 {self._sandbox.relative_to_root(dir_path)}/ 下未找到包含 '{pattern}' 的内容"

        header = f"搜索 '{pattern}' 找到 {len(matches)} 处匹配：\n"
        return header + "\n".join(matches)

    def _grep_search(
        self, dir_path: Path, regex: re.Pattern,
        matches: List[str], depth: int, max_depth: int,
    ) -> None:
        """递归搜索文件内容。"""
        if depth >= max_depth or len(matches) >= self._sandbox.max_results:
            return

        try:
            entries = sorted(dir_path.iterdir(), key=lambda e: e.name.lower())
        except PermissionError:
            return

        for entry in entries:
            if self._sandbox.is_excluded(entry):
                continue

            if len(matches) >= self._sandbox.max_results:
                matches.append(f"... 结果已达上限 ({self._sandbox.max_results})")
                return

            if entry.is_dir():
                self._grep_search(entry, regex, matches, depth + 1, max_depth)
            elif entry.is_file():
                self._search_in_file(entry, regex, matches)

    def _search_in_file(self, file_path: Path, regex: re.Pattern, matches: List[str]) -> None:
        """在单个文件中搜索匹配内容。"""
        # 跳过二进制文件
        if _is_binary_file(file_path):
            return
        # 跳过过大文件
        try:
            size = file_path.stat().st_size
            if size > self._sandbox.max_file_size:
                return
        except OSError:
            return

        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                rel_path = self._sandbox.relative_to_root(file_path)
                for line_no, line in enumerate(f, 1):
                    if len(matches) >= self._sandbox.max_results:
                        return
                    if regex.search(line):
                        # 截断过长的行
                        display_line = line.rstrip()
                        if len(display_line) > 200:
                            display_line = display_line[:200] + "..."
                        matches.append(f"  {rel_path}:{line_no}: {display_line}")
        except (UnicodeDecodeError, OSError):
            pass  # 跳过二进制文件或不可读文件

    def _read_file(self, path: str, offset: int = 0, limit: int = 0, **kwargs) -> str:
        """读取文件内容，支持分段读取。"""
        file_path = self._sandbox.validate_file_for_read(path)

        # Excel 文件使用专用解析器
        if file_path.suffix.lower() in _EXCEL_EXTENSIONS:
            return self._read_excel(file_path, offset=offset, limit=limit)

        # 其他二进制文件直接返回友好提示
        if _is_binary_file(file_path):
            return _binary_file_hint(file_path)

        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()

        total_lines = len(lines)
        rel_path = self._sandbox.relative_to_root(file_path)

        # 分段读取
        if offset > 0:
            start = max(0, offset - 1)  # 用户传的是 1-based 行号
        else:
            start = 0

        if limit > 0:
            end = min(start + limit, total_lines)
        else:
            end = total_lines

        selected = lines[start:end]

        # 添加行号
        numbered_lines = []
        for i, line in enumerate(selected, start + 1):
            numbered_lines.append(f"{i:>4}| {line.rstrip()}")

        content = "\n".join(numbered_lines)

        # 元信息头
        header = f"📄 {rel_path} (共 {total_lines} 行"
        if start > 0 or end < total_lines:
            header += f", 显示第 {start + 1}-{end} 行"
        header += ")\n"

        return header + content

    def _read_excel(self, file_path: Path, offset: int = 0, limit: int = 0) -> str:
        """使用 openpyxl 读取 Excel 文件，输出为 Markdown 表格。

        支持多 Sheet、分段读取（offset/limit 按数据行计算，不含表头）。
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return (
                "⚠️ 缺少 Excel 解析依赖，请执行: pip install openpyxl\n"
                f"文件: {file_path.name}"
            )

        rel_path = self._sandbox.relative_to_root(file_path)
        max_rows_per_sheet = 200  # 单个 Sheet 最大输出行数，防止超大文件爆 token

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            return f"⚠️ Excel 文件读取失败: {file_path.name}\n错误: {e}"

        try:
            sections: List[str] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                if not rows:
                    sections.append(f"### Sheet: {sheet_name}\n(空表)")
                    continue

                # 第一行作为表头
                headers = [str(c) if c is not None else "" for c in rows[0]]
                data_rows = rows[1:]
                total_data_rows = len(data_rows)

                # 分段读取
                start = max(0, offset - 1) if offset > 0 else 0
                end = min(start + limit, total_data_rows) if limit > 0 else total_data_rows
                # 截断保护
                if end - start > max_rows_per_sheet:
                    end = start + max_rows_per_sheet

                selected_rows = data_rows[start:end]

                # 构建 Markdown 表格
                col_count = len(headers)
                header_line = "| " + " | ".join(headers) + " |"
                separator = "| " + " | ".join(["---"] * col_count) + " |"

                table_lines = [header_line, separator]
                for row in selected_rows:
                    cells = []
                    for i in range(col_count):
                        val = row[i] if i < len(row) else None
                        cell_str = str(val) if val is not None else ""
                        # 转义 Markdown 管道符，截断过长单元格
                        cell_str = cell_str.replace("|", "\\|").replace("\n", " ")
                        if len(cell_str) > 100:
                            cell_str = cell_str[:100] + "..."
                        cells.append(cell_str)
                    table_lines.append("| " + " | ".join(cells) + " |")

                # Sheet 标题
                sheet_header = f"### Sheet: {sheet_name} ({total_data_rows} 行 × {col_count} 列"
                if start > 0 or end < total_data_rows:
                    sheet_header += f", 显示第 {start + 1}-{end} 行"
                if end - start >= max_rows_per_sheet and end < total_data_rows:
                    sheet_header += f", 已截断"
                sheet_header += ")"

                sections.append(sheet_header + "\n" + "\n".join(table_lines))

            file_header = f"📊 {rel_path} ({len(wb.sheetnames)} 个 Sheet)\n"
            return file_header + "\n\n".join(sections)
        finally:
            wb.close()
